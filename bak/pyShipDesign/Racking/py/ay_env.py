# acceleration parameter

from math import *
from openpyxl import *

global Ls, Cb, GM, g, fp

wb = load_workbook(filename = './Colorado_Summary.xlsx')
sheet_range = wb['Car Loading']


TLC = []
for i in range(2,14):
    TLC.append(sheet_range.cell(row = i, column = 9).value)
    print TLC
    
GM = []
for i in range(2,14):
    GM.append(sheet_range.cell(row = i, column = 10).value)
    print GM
    
mc = []
for i in range(0, 10):
    mc.append([])
    for j in range(33,43):
        mc[i].append(sheet_range.cell(row = j, column = i+1).value)
    print mc[i]

# for i in range(0,10):
#    print TLC[i]*GM[i]*mc[1][i]
#    print mc

# basic acceleration parameter - a0
def acceleration_parameter():

    a0 = (1.58 - 0.47 * Cb) * (2.4 / sqrt(Ls) + 34 / Ls - 600 / pow(Ls,2))

    return a0


# roll motion
def roll_motion():
    global g, kr, GM, fp, fBK
    
    # Roll period - Ttheta
    Ttheta = 2.3 * pi * kr / sqrt(g*GM);
    
    # Roll angle in deg. - theta
    theta = 9000 * (1.4 - 0.035 * Ttheta) * fp * fBK / ((1.15*B + 55)*pi)

    return Ttheta, theta

# sway acceleration
def sway_acceleration():
    global g, Ls, fp, a0
    
    a_sway = 0.3*(2.25-20/sqrt(g*Ls))*fp*a0*g

    return a_sway

# roll acceleration - aroll
# rad/sec^2
def roll_acceleration():
    global fp

    Ttheta, theta = roll_motion()
    
    a_roll = fp*theta* pi/180*pow(2* pi/Ttheta,2)
    
    return a_roll

# enveloped y acceleration
def envelope_transverse_accelerations(z):
    global Ls, GM, g
    
    R = min(D/4+TLC/2, D/2)
    Ttheta, theta = roll_motion()
    asway = sway_acceleration()
    aroll = roll_acceleration()
    
    aroll_y = aroll*(z-R);
    
    ay_env = (1- exp(-B*Ls/(215*GM)))*sqrt(asway^2 + (g*sind(theta)+aroll_y)^2)
        
    return ay_env

    
"""
Ls = 170.4
B = 30.2
Cb = 0.8
g = 9.81
kr = 11
GM = 1.4
a0 = acceleration_parameter()
fp = 1
fBK = 1
The, th = roll_motion()

print a0
print The, th
"""
