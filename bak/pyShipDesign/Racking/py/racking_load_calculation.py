# acceleration parameter

import openpyxl as xl
import math as mth
import xlsxwriter as xlwt
import xlrd

#basic acceleration parameter - a0
def acceleration_parameter():
    global Ls, Cb
    a0 = (1.58 - 0.47 * Cb) * (2.4 / mth.sqrt(Ls) + 34 / Ls - 600 / pow(Ls,2))

    return a0

# roll motion
def roll_motion(GM, kr):
    global B, g, fp, fBK
    # Roll period - Ttheta
    Ttheta = 2.3 * mth.pi * kr / mth.sqrt(g*GM)
    
    # Roll angle in deg. - theta
    theta = 9000 * (1.4 - 0.035 * Ttheta) * fp * fBK / ((1.15*B + 55)*mth.pi)

    return Ttheta, theta

# sway acceleration
def sway_acceleration():
    global Ls, g, a0, fp
    a_sway = 0.3*(2.25-20/mth.sqrt(g*Ls)) * fp * a0 * g

    return a_sway

# roll acceleration - aroll
# rad/sec^2
def roll_acceleration(GM, kr):
    global fp, fBK

    Ttheta, theta = roll_motion(GM, kr)
    
    a_roll = fp*theta* mth.pi/180*pow(2* mth.pi/Ttheta,2)
    
    return a_roll

# enveloped y acceleration
def envelope_transverse_accelerations(TLC, GM, kr, z):
    global Ls, D
    
    R = min(D/4+TLC/2, D/2)
    Ttheta, theta = roll_motion(GM, kr)
    asway = sway_acceleration()
    aroll = roll_acceleration(GM, kr)
    
    aroll_y = aroll*(z-R)
    
    ay_env = (1-mth.exp(-B*Ls/(215*GM)))*mth.sqrt(pow(asway,2)+pow(g*mth.sin(mth.radians(theta)) + aroll_y,2))
            
    return ay_env

def racking_moment(nDK, z_main, z, mc, ms, ay_env):
    MR = 0.0
    for i in range(0, nDK):
        MR = MR + (mc[i] + ms[i])*ay_env[i] *(z[i]-z_main)

    return MR

# row = 13, column = 12

wb = xl.load_workbook(filename = 'Colorado_Summary.xlsx')

sht = wb['Detail Of Car Loading']

TLC = []
for i in range(2,14):
    TLC.append(sht.cell(row = i, column = 8).value)
#    print TLC
    
GM = []
for i in range(2,14):
    GM.append(sht.cell(row = i, column = 9).value)
#    print GM

LWT = []
for i in range(2,14):
    LWT.append(sht.cell(row = i, column = 10).value)
#    print LWT

DWT = []
for i in range(2,14):
    DWT.append(sht.cell(row = i, column = 11).value)
#    print DWT

DSP = []
for i in range(2,14):
    DSP.append(sht.cell(row = i, column = 12).value)
#    print DSP

mc = []
for i in range(0, 12):
    mc.append([])
    for j in range(46, 33, -1):
        mc[i].append(sht.cell(row = j, column = i+2).value)
#    print mc[i]

ms = []
for i in range(63, 50, -1):
    ms.append(sht.cell(row = i, column = 3).value)
#    print ms

zi = []
for i in range(63, 50, -1):
    zi.append(sht.cell(row = i, column = 2).value)
#    print zi

# Main dimension
Ls = 170.4
B = 30.2
D = 28.8
Ts = 8.7
Cb = 0.56
Vs = 20.0
g = 9.81
a0 = acceleration_parameter()
fp = 1.0
fBK = 1.0
kr = 11.0
z_main = 14.4

# print acceleration_parameter()
# print sway_acceleration()
# print roll_acceleration(GM[0], kr)
# print roll_motion(GM[0], kr)
# print zi[0]

i = 0
for z in zi:
    i = i + 1
#    print i, z

j = 0    
for T in TLC:
    j = j + 1
#    print j, T

k = 0
for GMi in GM:
    k = k + 1
#    print k+1, GMi

for lci in range(0,12):
    ay_env = []
    for idk in range(0, 13):
        ay_env.append(envelope_transverse_accelerations(TLC[lci], GM[lci], kr, zi[idk]))
    
    MR = racking_moment(13, z_main, zi, mc[lci], ms, ay_env)
    print(ay_env)
    print(MR)

wbw = xlsw.workbook('Result_Load.xlsx')
shtw = wbw.add_worksheet()


print(25789/(1.025*170.4*30.2*8.7))
